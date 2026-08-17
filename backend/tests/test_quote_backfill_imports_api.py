from __future__ import annotations

import io
from collections.abc import Generator
from datetime import UTC, datetime
from typing import Any
from uuid import UUID, uuid4

import pytest
from fastapi import FastAPI
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook

from app.api.v1.quote_backfill_imports import (
    get_audit_log_service,
    get_file_admin_service,
    get_job_service,
)
from app.auth.dependencies import get_current_user
from app.db.session import get_db_session
from app.models import File, ImportJob, Permission, Role, User, UserStatus
from app.services.job_admin import JobNotFoundError


class MockSession:
    def __init__(self) -> None:
        self.committed = False

    async def commit(self) -> None:
        self.committed = True


class MockFileAdminService:
    async def upload_file(self, **kwargs: Any) -> File:
        return File(
            id=uuid4(),
            filename=kwargs["filename"],
            bucket="app-local",
            storage_path=f"path/{kwargs['filename']}",
            content_type=kwargs["content_type"],
            size_bytes=kwargs["size_bytes"],
            is_public=kwargs["is_public"],
            uploaded_by_id=kwargs["uploaded_by_id"],
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )


class MockJobAdminService:
    def __init__(self) -> None:
        self.session = MockSession()
        self.jobs: dict[UUID, ImportJob] = {}
        self.enqueue_observed_committed_states: list[bool] = []

    async def create_import_job(
        self,
        *,
        file_id: UUID,
        user_id: UUID,
        entity_type: str = "quote_backfill",
        task_name: str = "import_quote_backfill_task",
    ) -> ImportJob:
        db_file = File(
            id=file_id,
            filename="quote_backfill.csv",
            bucket="app-local",
            storage_path=f"path/{file_id}",
            content_type="text/csv",
            size_bytes=100,
            is_public=False,
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        job = ImportJob(
            id=uuid4(),
            file_id=file_id,
            entity_type=entity_type,
            task_name=task_name,
            status="pending",
            total_rows=0,
            processed_rows=0,
            failed_rows=0,
            created_by_id=user_id,
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        job.file = db_file
        self.jobs[job.id] = job
        return job

    async def enqueue_import_job(self, job: ImportJob) -> None:
        self.enqueue_observed_committed_states.append(self.session.committed)

    async def get_import_job_by_id(
        self,
        job_id: UUID,
        entity_type: str | None = None,
    ) -> ImportJob:
        job = self.jobs.get(job_id)
        if job is None or (entity_type is not None and job.entity_type != entity_type):
            raise JobNotFoundError()
        return job


class MockAuditLogService:
    def __init__(self) -> None:
        self.events: list[dict[str, object]] = []

    async def log_event(self, **kwargs: Any) -> None:
        self.events.append(kwargs)


def build_user(permission_codes: list[str]) -> User:
    user = User(id=uuid4(), email="admin@example.com", status=UserStatus.ACTIVE)
    role = Role(id=uuid4(), name="quote_backfill_importer", is_system=False)
    role.permissions = [Permission(id=uuid4(), code=code) for code in permission_codes]
    user.roles = [role]
    return user


@pytest.fixture
def override_dependencies(app: FastAPI) -> Generator[MockJobAdminService, None, None]:
    mock_job_service = MockJobAdminService()
    mock_audit_service = MockAuditLogService()

    app.dependency_overrides[get_current_user] = lambda: build_user(
        ["quotes.backfill_import"],
    )
    app.dependency_overrides[get_job_service] = lambda: mock_job_service
    app.dependency_overrides[get_file_admin_service] = lambda: MockFileAdminService()
    app.dependency_overrides[get_audit_log_service] = lambda: mock_audit_service
    app.dependency_overrides[get_db_session] = lambda: mock_job_service.session

    yield mock_job_service

    app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_download_quote_backfill_import_template_requires_permission(
    app: FastAPI,
    client: AsyncClient,
) -> None:
    app.dependency_overrides[get_current_user] = lambda: build_user([])

    denied_response = await client.get("/api/v1/quote-backfill-imports/template")
    app.dependency_overrides[get_current_user] = lambda: build_user(["quotes.backfill_import"])
    allowed_response = await client.get("/api/v1/quote-backfill-imports/template")

    assert denied_response.status_code == 403
    assert allowed_response.status_code == 200
    workbook = load_workbook(io.BytesIO(allowed_response.content), read_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    header_row = next(worksheet.iter_rows(values_only=True))
    assert "supplier_name" in header_row
    assert "import_tax_rate_percent" in header_row
    app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_start_quote_backfill_import_creates_job_and_enqueues_after_commit(
    client: AsyncClient,
    override_dependencies: MockJobAdminService,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(
        [
            "supplier_name",
            "received_date",
            "material_code",
            "price_original",
            "currency",
            "unit",
            "delivery_month",
            "exchange_rate",
            "import_tax_rate_percent",
            "processing_cost_vnd_per_kg",
            "note",
        ],
    )
    worksheet.append(
        [
            "TAN_LONG", "15/06/2026", "CORN", "300.00", "USD", "MT",
            "07/2026", "26100.00", "0.00", "200.00", None,
        ],
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    files = {
        "file": (
            "quote_backfill.xlsx",
            buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    response = await client.post("/api/v1/quote-backfill-imports", files=files)

    assert response.status_code == 201
    data = response.json()
    assert data["entity_type"] == "quote_backfill"
    job = override_dependencies.jobs[UUID(data["id"])]
    assert job.task_name == "import_quote_backfill_task"
    assert override_dependencies.enqueue_observed_committed_states == [True]


@pytest.mark.asyncio
async def test_start_quote_backfill_import_rejects_non_xlsx_file(
    client: AsyncClient,
    override_dependencies: MockJobAdminService,
) -> None:
    files = {"file": ("quote_backfill.csv", b"supplier_name\nTAN_LONG", "text/csv")}

    response = await client.post("/api/v1/quote-backfill-imports", files=files)

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_download_quote_backfill_import_error_file(
    client: AsyncClient,
    override_dependencies: MockJobAdminService,
) -> None:
    job = await override_dependencies.create_import_job(file_id=uuid4(), user_id=uuid4())
    job.status = "completed"
    job.failed_rows = 1
    job.errors_json = [{"row": 2, "errors": ["Nhà cung cấp 'BAD' không tồn tại."]}]

    response = await client.get(f"/api/v1/quote-backfill-imports/{job.id}/error-file")

    assert response.status_code == 200
    assert "BAD" in response.text
    assert "không tồn tại" in response.text


@pytest.mark.asyncio
async def test_get_quote_backfill_import_job_not_found(
    client: AsyncClient,
    override_dependencies: MockJobAdminService,
) -> None:
    response = await client.get(f"/api/v1/quote-backfill-imports/{uuid4()}")

    assert response.status_code == 404
