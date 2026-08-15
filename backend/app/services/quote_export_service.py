from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.sanitizer import clean_html_to_text

BUSINESS_TIMEZONE = ZoneInfo("Asia/Ho_Chi_Minh")

_TITLE = "BÁO CÁO BẢNG GIÁ NGUYÊN LIỆU"
_SHEET_NAME = "Bảng báo giá"

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=16, color="1F3864")
_SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
_BAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_STATUS_LABELS = {
    "draft": "Nháp",
    "confirmed": "Đã xác nhận",
    "superseded": "Đã bị thay thế",
}
_STATUS_FILLS = {
    "draft": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "confirmed": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    "superseded": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
}

# (row key, header label, column width, cell kind)
_COLUMNS: list[tuple[str, str, int, str]] = [
    ("received_date", "Ngày nhận", 14, "date"),
    ("supplier_name", "Nhà cung cấp", 26, "text"),
    ("material_name", "Vật tư", 22, "text"),
    ("price_original", "Giá gốc", 14, "number"),
    ("unit_label", "Đơn vị tính", 12, "text_center"),
    ("exchange_rate", "Tỷ giá", 12, "number"),
    ("delivery_month", "Kỳ giao hàng", 14, "month"),
    ("import_tax_rate_percent", "Thuế NK (%)", 12, "percent"),
    ("processing_cost_vnd_per_kg", "Chi phí làm hàng (VNĐ/KG)", 20, "number"),
    ("price_converted_vnd_per_kg", "Giá quy đổi (VNĐ/KG)", 20, "number_bold"),
    ("purchased_label", "Chốt mua", 12, "text_center"),
    ("version_status", "Trạng thái", 16, "status"),
    ("note_content", "Ghi chú", 36, "text_wrap"),
    ("note_author_name", "Người ghi chú", 18, "text"),
    ("note_created_at", "Thời gian ghi chú", 17, "datetime"),
]

_HEADER_ROW = 3
_FIRST_DATA_ROW = 4


def _to_local(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    return value.astimezone(BUSINESS_TIMEZONE).replace(tzinfo=None)


def _to_float(value: Decimal | float | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _apply_data_cell_style(cell: Any, kind: str, is_band_row: bool) -> None:
    cell.border = _THIN_BORDER
    if is_band_row:
        cell.fill = _BAND_FILL

    if kind in ("date", "month", "datetime", "text_center"):
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif kind == "text_wrap":
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    elif kind in ("number", "number_bold", "percent"):
        cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    if kind == "date":
        cell.number_format = "dd/mm/yyyy"
    elif kind == "month":
        # "Kỳ giao hàng" chỉ có ý nghĩa theo THÁNG (luôn là ngày 1 của tháng
        # trong DB) — hiện dd/mm/yyyy như received_date sẽ gây hiểu nhầm có
        # ngày cụ thể trong tháng.
        cell.number_format = "mm/yyyy"
    elif kind == "datetime":
        cell.number_format = "dd/mm/yyyy hh:mm"
    elif kind == "percent":
        cell.number_format = '0.00"%"'
    elif kind in ("number", "number_bold"):
        cell.number_format = "#,##0.00"

    if kind == "number_bold":
        cell.font = Font(bold=True, color="1F3864")


def _cell_value(row: dict[str, Any], key: str, kind: str) -> Any:
    if key == "unit_label":
        currency = row.get("currency") or ""
        unit = row.get("unit") or ""
        return f"{currency}/{unit}" if currency or unit else None
    if key == "purchased_label":
        return "Đã chốt mua" if row.get("purchased") else "Chưa chốt"
    if key == "version_status":
        status = row.get("version_status")
        if not isinstance(status, str):
            return status
        return _STATUS_LABELS.get(status, status)
    if key == "note_content":
        # Ghi chú lưu dạng HTML (soạn qua trình soạn thảo rich-text ở
        # frontend) — chuyển về văn bản thuần trước khi ghi vào Excel, tránh
        # lộ thẻ HTML/entity thô ("<p>", "&nbsp;") trong báo cáo.
        return clean_html_to_text(row.get("note_content"))

    value = row.get(key)
    if kind == "datetime":
        return _to_local(value)
    if kind in ("number", "number_bold", "percent"):
        return _to_float(value)
    return value


def _write_title_rows(
    sheet: Worksheet,
    *,
    last_col_letter: str,
    generated_at: datetime,
    row_count: int,
) -> None:
    sheet.merge_cells(f"A1:{last_col_letter}1")
    title_cell = sheet["A1"]
    title_cell.value = _TITLE
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26

    sheet.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = sheet["A2"]
    local_generated_at = generated_at.astimezone(BUSINESS_TIMEZONE)
    subtitle_cell.value = (
        f"Ngày xuất báo cáo: {local_generated_at.strftime('%d/%m/%Y %H:%M')}"
        f"   |   Tổng số dòng: {row_count}"
    )
    subtitle_cell.font = _SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[2].height = 18


def _write_header_row(sheet: Worksheet) -> None:
    for col_index, (_, label, width, _) in enumerate(_COLUMNS, start=1):
        cell = sheet.cell(row=_HEADER_ROW, column=col_index, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    sheet.row_dimensions[_HEADER_ROW].height = 20


def _write_data_rows(sheet: Worksheet, rows: list[dict[str, Any]]) -> None:
    for row_offset, row in enumerate(rows):
        excel_row = _FIRST_DATA_ROW + row_offset
        is_band_row = row_offset % 2 == 1
        for col_index, (key, _, _, kind) in enumerate(_COLUMNS, start=1):
            cell = sheet.cell(
                row=excel_row,
                column=col_index,
                value=_cell_value(row, key, kind),
            )
            _apply_data_cell_style(cell, kind, is_band_row)

            if key == "version_status":
                status_value = row.get("version_status")
                status_fill = (
                    _STATUS_FILLS.get(status_value) if isinstance(status_value, str) else None
                )
                if status_fill:
                    cell.fill = status_fill


def build_quote_export_workbook(
    rows: list[dict[str, Any]],
    generated_at: datetime | None = None,
) -> bytes:
    generated_at = generated_at or datetime.now(BUSINESS_TIMEZONE)

    workbook = Workbook()
    active_sheet = workbook.active
    assert isinstance(active_sheet, Worksheet)  # noqa: S101 — a fresh Workbook() always has one
    sheet = active_sheet
    sheet.title = _SHEET_NAME

    last_col_letter = get_column_letter(len(_COLUMNS))

    _write_title_rows(
        sheet,
        last_col_letter=last_col_letter,
        generated_at=generated_at,
        row_count=len(rows),
    )
    _write_header_row(sheet)
    _write_data_rows(sheet, rows)

    last_row = _FIRST_DATA_ROW + max(len(rows), 1) - 1
    sheet.auto_filter.ref = f"A{_HEADER_ROW}:{last_col_letter}{last_row}"
    sheet.freeze_panes = f"A{_FIRST_DATA_ROW}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
